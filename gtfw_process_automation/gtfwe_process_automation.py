# Copy this script under function directiory:
# It execute script, generate log, analyse data and generate excel report

import os, sys, re, xlsxwriter
import openpyxl

import jedi

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

from datetime import datetime
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle


#==================================================================================================



#===================================================================================================================================
# Execute test scripts in the given order
# datetime object containing current date and time
now = datetime.now()
dt_string = now.strftime("%d_%m_%Y_%H_%M_%S")
print("date and time =", dt_string)

# Get the absolute path
dir_path = os.path.dirname(os.path.abspath(__file__))
#print("0000 : "+dir_path)

# Directories name
analysis_dir = dir_path+"/"+"Analysis"
logs_dir = analysis_dir+"/"+"Execution_Logs"
reports_dir = analysis_dir+"/"+"Reports"
indiv_report_dir = reports_dir+"/"+"Indivisual_Reports"
collect_report_dir = reports_dir+"/"+"Collective_Reports"

# Create directories if doesn't exist
if not os.path.exists(analysis_dir):
    os.makedirs(analysis_dir)
if os.path.exists(analysis_dir):
    os.chdir(analysis_dir)
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
    if os.path.exists(reports_dir):
        os.chdir(reports_dir)
        if not os.path.exists(indiv_report_dir):
            os.makedirs(indiv_report_dir)
        if not os.path.exists(collect_report_dir):
            os.makedirs(collect_report_dir)

print("Execution is going to start!!!")

# Create log file to save test scripts execution output
LOG = logs_dir+"/gtfwe_log_"+dt_string+".log"
sys.stdout = open(LOG, 'w')


## Execute all the scripts exists under function/WiFi directory
#for name in os.listdir(dir_path):
#    if os.path.isdir(os.path.join(dir_path, name)):
#        if (name.startswith("WiFi")):
#            os.chdir(os.path.join(dir_path, name))
#            for name in os.listdir(dir_path+"/"+name):
#                if (name.startswith("WiFi")) and name.endswith(".py"):
#                    print("Testscripts: "+str(name)+"")
#                    print(os.popen("python3 "+str(name)+"").read())
#


# Execute all the scripts exists under function directory
for name in os.listdir(dir_path):
    if os.path.isdir(os.path.join(dir_path, name)):
        if (name.startswith("WiFi") or name.startswith("WAN") or name.startswith("IPTV") or name.startswith("Network") or name.startswith("System-OS")):        
            os.chdir(os.path.join(dir_path, name))
            for name in os.listdir(dir_path+"/"+name):
                if (name.startswith("WiFi") or name.startswith("WAN") or name.startswith("IPTV") or name.startswith("SDK") or name.startswith("Network") or name.startswith("System-OS") or name.startswith("Security")) and name.endswith(".py") and ("Network_1105_UPNP_functionality.py" not in name and "Network_968_dhcp_option_252.py" not in name and "Network_969_DHCP_static_leases.py" not in name and "Network_1086_dhcp_resolve_domain_name.py" not in name):
                    print("Testscripts: "+str(name)+"")
                    print(os.popen("python3 "+str(name)+"").read())


sys.stdout.close()
#===================================================================================================================================

# Analyse data from generated log 
with open(LOG) as f:
    lines = f.readlines()

tescript_names = []
feature = []
results = []
failure_reasons = []
fail_res_update_done = True
sheet_name = None
hardware = None
build = None

def escape_ansi(line):
    ansi_escape = re.compile(r'(\x9B|\x1B\[)[0-?]*[ -/]*[@-~]')
    return ansi_escape.sub('', line)


for line in lines:
    line = escape_ansi(line)
    if  'Testscripts: ' in line and fail_res_update_done:
        words = line.split(' ',1)
        ts_names = words[1].rstrip()
        tescript_names.append(ts_names)
        feat = words[1].split('_',1)
        feature.append(feat[0])
        continue
    if 'TAP version 13' in line:
        continue
    if ('ok ' or 'not ok ') in line:
        words = line.split('1',1)
        ts_results = words[0].rstrip()
        if ts_results == 'ok':
            ts_results = 'PASS'
            ts_reason = ''
            results.append(ts_results)
            failure_reasons.append(ts_reason)
            fail_res_update_done = True
            continue
        elif ts_results == 'not ok':
            ts_results = 'FAIL'
            results.append(ts_results)
            fail_res_update_done = False
            continue
    if 'Error:' in line and 'Device' not in line and 'Error: not' not in line:
         ts_reason = line.strip()
         failure_reasons.append(ts_reason)
         fail_res_update_done = True
         continue
    if 'firmware version:' in line:
        words = line.split('firmware version:')
        build = words[1].strip()
        sheet_title = build.split('-')
        sht_title = []
        hardware = sheet_title[0]
        sht_title.append(sheet_title[0])
        sht_title.append(sheet_title[2])
        sht_title.append(sheet_title[3])
        sht_title.append(sheet_title[4])

        sheet_name = "_"
        build = sheet_name.join(sht_title)


#===================================================================================================================================
# Create Indivisual status excel to enter analysed data
REPORT = indiv_report_dir+"/gtfwe_status_"+dt_string+".xlsx"
workbook = xlsxwriter.Workbook(REPORT)
worksheet = workbook.add_worksheet('Status')

cell_format_headings = workbook.add_format({'bold': True,
                                            'align': 'center',
                                            'fg_color': '#808080',
                                            'font_name': 'Times New Roman',
                                            'font_size' : 16,
                                            'font_color' : 'black',
                                            'text_wrap': True,
                                            'border': 5 })
cell_format_content = workbook.add_format({ 'align': 'left',
                                            'font_name': 'Times New Roman',
                                            'font_size' : 12,
                                            'font_color' : 'black',
                                            'text_wrap': True,
                                            'border': 1 })

worksheet.write('A1', 'Testscripts', cell_format_headings)
worksheet.write('B1', 'Feature', cell_format_headings)
worksheet.write('C1', 'Hardware', cell_format_headings)
worksheet.write('D1', 'Build', cell_format_headings)
worksheet.write('E1', 'Execution Status', cell_format_headings)
worksheet.write('F1', 'Failure Reason', cell_format_headings)
worksheet.write('G1', 'Comments', cell_format_headings)
worksheet.autofilter('A1:E1')
worksheet.set_tab_color('orange')
worksheet.freeze_panes('A2')

worksheet.set_column(0, 0, 60)
worksheet.set_column(1, 1, 15)
worksheet.set_column(2, 2, 18)
worksheet.set_column(3, 3, 50)
worksheet.set_column(4, 4, 15)
worksheet.set_column(5, 5, 100)
worksheet.set_column(6, 6, 60)
worksheet.set_row(0, 40)

red_format = workbook.add_format({'align': 'center',
                                  'font_color': 'red'})

green_format = workbook.add_format({'align': 'center',
                                    'font_color': 'green'})

row = 1
column = 0
for i in range(len(tescript_names)):
    worksheet.write(row, column, str(tescript_names[i]), cell_format_content)
    row = row + 1


row = 1
column = 1
for i in range(len(feature)):
    worksheet.write(row, column, str(feature[i]), cell_format_content)
    row = row + 1


row = 1
column = 2
for i in range(len(tescript_names)):
    worksheet.write(row, column, str(hardware), cell_format_content)
    row = row + 1

row = 1
column = 3
for i in range(len(tescript_names)):
    worksheet.write(row, column, str(build), cell_format_content)
    row = row + 1

row = 1
column = 4
for i in range(len(results)):
    worksheet.write(row, column, str(results[i]), cell_format_content)
    worksheet.conditional_format(4, 5, row, 2, {
                                                     'type': 'cell',
                                                     'criteria': 'equal to',
                                                     'value': '"FAIL"',
                                                     'format': red_format
                                                    })
    worksheet.conditional_format(4, 5, row, 2, {
                                                     'type': 'cell',
                                                     'criteria': 'equal to',
                                                     'value': '"PASS"',
                                                     'format': green_format
                                                    })
    row = row + 1

row = 1
column = 5
for i in range(len(failure_reasons)):
    worksheet.write(row, column, str(failure_reasons[i]), cell_format_content)
    row = row + 1

workbook.close()

#===================================================================================================================================

## Create collective status excel to enter analysed data in appended format, if doesnt exist
COLLECTIVE_REPORT = collect_report_dir+"/gtfwe_collective_status.xlsx"
if not os.path.isfile(COLLECTIVE_REPORT):
    cworkbook = xlsxwriter.Workbook(COLLECTIVE_REPORT)

    cworksheet = cworkbook.add_worksheet('Status')
    
    cell_format_headings = cworkbook.add_format({'bold': True,
                                            'align': 'center',
                                            'fg_color': '#808080',
                                            'font_name': 'Times New Roman',
                                            'font_size' : 16,
                                            'font_color' : 'black',
                                            'text_wrap': True,
                                            'border': 5 })
    cell_format_content = cworkbook.add_format({ 'align': 'left',
                                            'font_name': 'Times New Roman',
                                            'font_size' : 12,
                                            'font_color' : 'black',
                                            'text_wrap': True,
                                            'border': 1 })

    cworksheet.write('A1', 'Testscripts', cell_format_headings)
    cworksheet.write('B1', 'Feature', cell_format_headings)
    cworksheet.write('C1', 'Hardware', cell_format_headings)
    cworksheet.write('D1', 'Build', cell_format_headings)
    cworksheet.write('E1', 'Execution Status', cell_format_headings)
    cworksheet.write('F1', 'Failure Reason', cell_format_headings)
    cworksheet.write('G1', 'Comments', cell_format_headings)
    cworksheet.autofilter('A1:E1')
    cworksheet.set_tab_color('orange')
    cworksheet.freeze_panes('A2')

    cworksheet.set_column(0, 0, 60)
    cworksheet.set_column(1, 1, 15)
    cworksheet.set_column(2, 2, 18)
    cworksheet.set_column(3, 3, 50)
    cworksheet.set_column(4, 4, 15)
    cworksheet.set_column(5, 5, 100)
    cworksheet.set_column(6, 6, 60)
    cworksheet.set_row(0, 40)

    cworkbook.close()

if os.path.isfile(COLLECTIVE_REPORT):
    cworkbook = openpyxl.load_workbook(COLLECTIVE_REPORT)
    cworksheet = cworkbook.active

    big_red_text = Font(color=colors.RED, size=14)
    center_aligned_text = Alignment(horizontal="center")

    max_row = cworksheet.max_row 
    row = max_row + 1 
    column = 1
    for i in range(len(tescript_names)):
        cworksheet.cell(row, column, str(tescript_names[i]))
        row = row + 1

    row = max_row + 1 
    column = 2
    for i in range(len(feature)):
        cworksheet.cell(row, column, str(feature[i]))
        row = row + 1


    row = max_row + 1 
    column = 3
    for i in range(len(tescript_names)):
        cworksheet.cell(row, column, str(hardware))
        row = row + 1

    row = max_row + 1 
    column = 4
    for i in range(len(tescript_names)):
        cworksheet.cell(row, column, str(build))
        row = row + 1

    row = max_row + 1 
    column = 5
    for i in range(len(results)):
        cworksheet.cell(row, column, str(results[i]))
        row = row + 1


    row = max_row + 1 
    column = 6
    for i in range(len(failure_reasons)):
        cworksheet.cell(row, column, str(failure_reasons[i]))
        row = row + 1


    cworkbook.save(COLLECTIVE_REPORT)
    cworkbook.close()

#==================================================================================================

email_user = 'latha.arun@iopsys.eu'
email_password = '$GirijaSangaiah123'
recipients = ['latha.arun@iopsys.eu']

subject = 'Test_python_attachment'
msg = MIMEMultipart()
msg['From'] = email_user
msg['To'] = ", ".join(recipients)
msg['Subject'] = subject

#email content
message = """<html>
<body>
Please find the attachment as test scripts execution log, execution status excel report.
<br><br>

Let me know if you have any questions

</body>
</html>
"""

msg.attach(MIMEText(message, 'html'))


files = [ LOG,
          REPORT,
          COLLECTIVE_REPORT ]

for a_file in files:
    attachment = open(a_file, 'rb')
    file_name = os.path.basename(a_file)
    part = MIMEBase('application','octet-stream')
    part.set_payload(attachment.read())
    part.add_header('Content-Disposition',
                    'attachment',
                    filename=file_name)
    encoders.encode_base64(part)
    msg.attach(part)


text = msg.as_string()
server = smtplib.SMTP('smtp.office365.com',587)
server.starttls()
server.login(email_user,email_password)

server.sendmail(email_user,recipients,text)
server.quit()




